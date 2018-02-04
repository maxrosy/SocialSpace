import pandas as pd
import requests
import sys
import boto3
import botocore
import datetime
import time
import configparser
from openpyxl import load_workbook, Workbook
from sqlalchemy import create_engine, MetaData,Table
from sqlalchemy.dialects.mysql import insert
from sqlalchemy.types import *
from SocialAPI.Logger.BasicLogger import Logger
import re
from ..Model import engine
from sqlalchemy.orm import sessionmaker
from SocialAPI.Helper import Helper
import aiohttp



class SocialBasicAPI(object):
	
	def __init__(self):
		self.__rootPath = Helper().getRootPath()
		self.cfp = configparser.ConfigParser()
		self.cfp.read(self.__rootPath + '/conf/social.conf')
		self.logger = Logger(self.__rootPath + '/conf/logging.conf','simpleExample').createLogger()
		self.__username = self.cfp.get('db','user')
		self.__password = self.cfp.get('db','password')
		self.__host = self.cfp.get('db','host')
		self.__port = int(self.cfp.get('db','port'))
		#self.__db = self.cfp.get('db','db')
		#self.__table = self.cfp.get('db','table')
		
	def postRequest(self,url, postData):
		
		r = requests.post(url, data=postData)
		return r
		
	def getRequest(self,url,paramsDict={},stream=False):
		
		r = requests.get(url, params=paramsDict, stream=stream)
		return r

	async def getAsyncRequest(self,url, paramsDict={}):
		async with aiohttp.ClientSession() as session:
			async with session.get(url, params=paramsDict) as r:
				return  await r.json()

	def encodeElement(self,text):

		if isinstance(text,dict) or isinstance(text,list):
			return str(text).encode('utf8')
		else:
			return text

	def cleanRecords(self,df,dropColumns=[],dedupColumns=[],renameColumns={},utcTimeCovert=True):
		self.logger.info("Calling cleanRecords function")
		try:
			len_before_etl = len(df)

			if dropColumns:
				df.drop(columns=dropColumns, inplace=True)

			# Change [] to None before encoding
			df = df.where(df.applymap(lambda x: False if not x else True), None)

			df=df.applymap(self.encodeElement)
			df.drop_duplicates(inplace=True)

			if dedupColumns:
				isDuplicated = df.duplicated(dedupColumns)
				df = df[~isDuplicated]
			if renameColumns:
				df.rename(columns=renameColumns, inplace=True)
			if utcTimeCovert:
				df['created_at'] = df['created_at'].apply(lambda x: datetime.datetime.strptime(x,"%a %b %d %H:%M:%S %z %Y"))
			df.reset_index(drop=True,inplace=True)

			# Change NaN to None
			df = df.where(pd.notnull(df),None)


			len_after_etl = len(df)
			self.logger.info('Totally {} out of {} records remained after ETL'.format(len_after_etl, len_before_etl))
			return df
		except Exception as e:
			self.logger.error('On line {} - {}'.format(sys.exc_info()[2].tb_lineno,e))
			exit(1)
	
	def writeDataFrameToCsv(self,df,filename,sep=','):
		self.logger.info("Calling writeDataFrameToCsv function")
		try:
			df.to_csv(filename,sep=sep,header=True,index=False)
		except Exception as e:
			self.logger.error('On line {} - {}'.format(sys.exc_info()[2].tb_lineno,e))
			exit(1)	

	def writeDataFrameToExcel(self,df,wbname,sheetname):
		self.logger.info("Calling writeDataFrameToExcel function")

		try:
			try:
				writer = pd.ExcelWriter(wbname)
				wb = load_workbook(wbname)
					
			except FileNotFoundError:
				self.logger.warn("File {} is not found! Creating one".format(wbname))
				wb = Workbook()
				
			writer.book = wb
			writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
					
			df.to_excel(writer,sheetname,index=False)
			writer.save()
		
		except Exception as e:
			self.logger.error('On line {} - {}'.format(sys.exc_info()[2].tb_lineno,e))
			exit(1)
			
		"""
		try:
			try:
				wb = load_workbook(wbname)
				ws = wb.create_sheet(sheetname, 0)

			except FileNotFoundError:
				self.logger.warn("File {} is not found! Creating one".format(wbname))
				wb = Workbook()
				ws = wb.active
				ws.title = sheetname

			df = df.replace([True,False],[1,0],inplace=True)
			for row in dataframe_to_rows(df, index=False, header=True):
				ws.append(row)
			wb.save(wbname)
		
		except Exception as e:
			self.logger.error('On line {} - {}'.format(sys.exc_info()[2].tb_lineno,e))
			exit(1)
		"""

	def readCsvToDataFrame(filePath,sep=','):	
		self.logger.info("Calling readCsvToDataFrame function")
		try:
			reader = pd.read_csv(filePath,sep=sep,iterator=True)
			loop = True
			chunkSize = 1000
			chunks = []
			n = 1
			while loop:
				try:
					chunk = reader.get_chunk(chunkSize)
					chunks.append(chunk)
					n += 1
				except StopIteration:
					loop = False
					self.logger.info("Iteration is stopped. Totally, {} loop(s)".format(n))
			df = pd.concat(chunks, ignore_index=True)
			self.logger.info("Totally {} records imported".format(len(df)))
			#df.info()
			return df
		except Exception as e:
			self.logger.error('On line {} - {}'.format(sys.exc_info()[2].tb_lineno,e))
			exit(1)
			
	def syncToDBUsingPandas(self,df,db,table,syncType='append'):
		"""
		Call this function with caution, it will overwrite table schema, if syncType is replace
		:param df:
		:param db:
		:param table:
		:param syncType: append/replace
		:return:
		"""
		self.logger.info("Calling syncToDBUsingPandas function")
		try:
			#df['created_time'] = datetime.datetime.now()
			#df['updated_time'] = datetime.datetime.now()
			dblink = 'mysql+mysqldb://{}:{}@{}/{}?charset=utf8'.format(self.__username,self.__password,self.__host,db)
			engine = create_engine(dblink,encoding='utf-8')
			#df.to_sql(table,engine,chunksize=1000,dtype={"Agency": String(50),"Platform":String(50),"Likes":Integer},index=False,if_exists='append',encoding='utf-8')
			df.to_sql(table,engine,chunksize=1000,index=False,if_exists=syncType)

		except Exception as e:
			self.logger.error('On line {} - {}'.format(sys.exc_info()[2].tb_lineno,e))
			exit(1)

	def readFromDBUsingPandas(self, db, table):
		self.logger.info("Calling readFromDBUsingPandas")
		try:
			dblink = 'mysql+mysqldb://{}:{}@{}/{}?charset=utf8'.format(self.__username, self.__password, self.__host,db)
			engine = create_engine(dblink, encoding='utf-8')
			records = pd.read_sql_table(table,engine)

			return records

		except Exception as e:
			self.logger.error('On line {} - {}'.format(sys.exc_info()[2].tb_lineno, e))
			exit(1)

	def createSession(self, engine=engine):
		Session = sessionmaker(bind=engine)
		session = Session()
		return session

	def upsertToDB(self,table,records):
		"""
		New feature in SQLAlchemy 1.2
		:param table: table model
		:param records: either in the type of dict or dataframe
		:return:
		"""
		try:
			#engine, meta = self.connectToDB(dbName)
			conn = engine.connect()
			#table = Table(tableName, meta)

			# Check data type of records
			if isinstance(records, pd.DataFrame):
				records = records.to_dict('records')
			elif isinstance(records, dict):
				pass
			else:
				raise Exception("Record Type {} is wrong".format(type(records)))

			# To do, check if batch upsert is possible
			for record in records:
				insert_stmt = insert(table).values(record)
				#record.pop(pk)
				upsert_stmt = insert_stmt.on_duplicate_key_update(**record)
				res = conn.execute(upsert_stmt)

			res.close()
			conn.close()
			self.logger.info("{} reocrds have been upsert into table {}".format(len(records),table.__table__))
		except Exception as e:
			self.logger.error('On line {} - {}'.format(sys.exc_info()[2].tb_lineno, e))
			exit(1)

	def insertToDB(self,dbName,tableName,records):

		try:
			engine, meta = self.connectToDB(dbName)
			conn = engine.connect()

			table = Table(tableName,meta)
			stmt = table.insert()

			# Check data type of records
			if isinstance(records, pd.DataFrame):
				res = conn.execute(stmt,records.to_dict('records'))
			elif isinstance(records,dict):
				res = conn.execute(stmt,records)
			else:
				raise Exception("Record Type {} is wrong".format(type(records)))
				
			self.logger.info('{} record(s) have been inserted into {}'.format(res.rowcount,tableName))

			res.close()
			conn.close()
		except Exception as e:
			self.logger.error('On line {} - {}'.format(sys.exc_info()[2].tb_lineno,e))
			exit(1)

	def connectToDB(self, db):
		
		try:
			dblink = 'mysql+mysqldb://{}:{}@{}/{}?charset=utf8mb4'.format(self.__username,self.__password,self.__host,db)
			engine = create_engine(dblink,echo=False)
			meta = MetaData(bind=engine,reflect=True)
			
			return (engine, meta)
		
		except Exception as e:
			self.logger.error('On line {} - {}'.format(sys.exc_info()[2].tb_lineno,e))
			exit(1)

	def readFromS3(self,bucketName,remoteFile,localFile):
		self.logger.info("Calling readFromS3 function")
		try:
			s3 = boto3.resource('s3')
			s3.Bucket(bucketName).download_file(remoteFile, localFile)
		
		except botocore.exceptions.ClientError as e:
			if e.response['Error']['Code'] == "404":
				self.logger.error("The object does not exist.")
			else:
				raise
		
		except Exception as e:
			self.logger.error('On line {} - {}'.format(sys.exc_info()[2].tb_lineno,e))
			exit(1)
				
	def writeToS3(self,bucketName,localFile,remoteFile):
		self.logger.info("Calling writeToS3 function")
		try:
			s3 = boto3.resource('s3')
			data = open(localFile, 'rb')
			s3.Bucket(bucketName).put_object(Key=remoteFile, Body=data)
			
		except Exception as e:
			self.logger.error('On line {} - {}'.format(sys.exc_info()[2].tb_lineno,e))
			exit(1)

	def getStrTime(self, delta=0):
		date_time = datetime.date.today() - datetime.timedelta(days=delta)

		return date_time.strftime("%Y-%m-%d %H:%M:%S")

	def getTimeStamp(self, datestr,level='s'):
		date_time = datetime.datetime.strptime(datestr, '%Y-%m-%d %H:%M:%S')
		if level == 's':
			return int(time.mktime(date_time.timetuple()))
		elif level =='ms':
			return round(time.mktime(date_time.timetuple())*1000)
		else:
			raise Exception('Time level is either s or ms')

	def matchPostSource(self,text):
		matchObj = re.search(r'^<a.*>(.*)</a>$', text)
		if matchObj:
			return matchObj.group(1)
		else:
			return text

	def __str__(self):
		return "Basic API of Social"
		
